import os
import shutil
import zipfile
import json
from datetime import datetime , timedelta
from fastapi import Body
import random
from fastapi_mail import FastMail, MessageSchema, ConnectionConfig
from fastapi import FastAPI, File, UploadFile, Form, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session , joinedload
from io import BytesIO
from fastapi.responses import StreamingResponse
from uuid import uuid4
from database import engine, SessionLocal
import models
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from passlib.context import CryptContext
pwd_context = CryptContext(schemes=["argon2"], deprecated="auto")

# =====================================================
# APP SETUP
# =====================================================

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,                                              
    allow_methods=["*"],
    allow_headers=["*"],
)

models.Base.metadata.create_all(bind=engine)


conf = ConnectionConfig(
    MAIL_USERNAME="careerset.notification@gmail.com",
    MAIL_PASSWORD="qezw emaa dwxv lgmy",  # no spaces
    MAIL_FROM="careerset.notification@gmail.com",
    MAIL_PORT=587,
    MAIL_SERVER="smtp.gmail.com",
    MAIL_STARTTLS=True,
    MAIL_SSL_TLS=False,
    USE_CREDENTIALS=True
)


# =====================================================
# UPLOAD FOLDER SETUP
# =====================================================

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")


# =====================================================
# DATABASE SESSION
# =====================================================

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# =====================================================
# FILE SAVE FUNCTION
# =====================================================

def save_file(file: UploadFile):
    if not file:
        return None

    file_path = os.path.join(UPLOAD_FOLDER, file.filename)

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return file.filename




# =====================================================
# UPLOAD FOLDER SETUP
# =====================================================

UPLOAD_FOLDER = os.path.abspath("uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

# =====================================================
# DATABASE SESSION
# =====================================================

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# =====================================================
# FILE SAVE FUNCTION
# =====================================================

def save_file(file: UploadFile):
    if not file:
        return None

    unique_name = f"{uuid4()}_{file.filename}"
    file_path = os.path.join(UPLOAD_FOLDER, unique_name)

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return unique_name

# =====================================================
# EMPLOYEE SUBMISSION
# =====================================================

@app.post("/employee-joining")
async def employee_joining(

    # BASIC
    employeeCode: str = Form(None),
    name: str = Form(...),
    dob: str = Form(...),
    gender: str = Form(...),
    phone: str = Form(...),
    email: str = Form(...),
    doj: str = Form(None),

    fatherName: str = Form(None),
    motherName: str = Form(None),
    department: str = Form(None),
    designation: str = Form(None),
    bloodGroup: str = Form(None),
    maritalStatus: str = Form(None),
    spouseName: str = Form(None),
    panNumber: str = Form(None),
    aadharNumber: str = Form(None),
    permanentAddress: str = Form(None),
    presentAddress: str = Form(None),

    # EMERGENCY
    emergencyName: str = Form(None),
    emergencyRelation: str = Form(None),
    emergencyPhone: str = Form(None),

    # EDUCATION
    qualification10: str = Form(None),
    year10: str = Form(None),
    percent10: str = Form(None),

    qualification12: str = Form(None),
    year12: str = Form(None),
    percent12: str = Form(None),

    ugDegree: str = Form(None),
    ugCollege: str = Form(None),
    ugYear: str = Form(None),
    ugPercent: str = Form(None),

    pgDegree: str = Form(None),
    pgCollege: str = Form(None),
    pgYear: str = Form(None),
    pgPercent: str = Form(None),

    # EXPERIENCE
    totalExpYears: str = Form(None),
    totalExpMonths: str = Form(None),
    careerBreak: str = Form(None),
    careerBreakDuration: str = Form(None),
    careerBreakReason: str = Form(None),

    # BANK
    accountHolderName: str = Form(None),
    bankName: str = Form(None),
    accountNumber: str = Form(None),
    ifscCode: str = Form(None),
    branchName: str = Form(None),

    # ESI / PF
    esiApplicable: str = Form(None),
    uanNumber: str = Form(None),
    pfNumber: str = Form(None),
    esiNumber: str = Form(None),

    # RELATIONS JSON
    dependents: str = Form(None),
    trainings: str = Form(None),
    employments: str = Form(None),

    # DOCUMENT FILES
    resume: UploadFile = File(None),
    sslc: UploadFile = File(None),
    hsc: UploadFile = File(None),
    aadharSelf: UploadFile = File(None),
    photo: UploadFile = File(None),
    aadharFather: UploadFile = File(None),
    aadharMother: UploadFile = File(None),
    panSelf: UploadFile = File(None),
    bankPassbookPhoto: UploadFile = File(None),

    db: Session = Depends(get_db)
):

    dob_converted = datetime.strptime(dob, "%Y-%m-%d").date()
    doj_converted = datetime.strptime(doj, "%Y-%m-%d").date() if doj else None
    # DEPENDENT FILES (Dynamic)
    dependent_aadhar_0: UploadFile = File(None),
    dependent_pan_0: UploadFile = File(None),
    dependent_photo_0: UploadFile = File(None),
    # Save Files
    resume_file = save_file(resume)
    sslc_file = save_file(sslc)
    hsc_file = save_file(hsc)
    aadhar_self_file = save_file(aadharSelf)
    photo_file = save_file(photo)
    aadhar_father_file = save_file(aadharFather)
    aadhar_mother_file = save_file(aadharMother)
    pan_self_file = save_file(panSelf)
    bank_passbook_file = save_file(bankPassbookPhoto)

    employee = models.EmployeeJoining(
        employee_code=employeeCode,
        name=name,
        dob=dob_converted,
        gender=gender,
        phone=phone,
        email=email,
        doj=doj_converted,

        father_name=fatherName,
        mother_name=motherName,
        department=department,
        designation=designation,
        blood_group=bloodGroup,
        marital_status=maritalStatus,
        spouse_name=spouseName,
        pan_number=panNumber,
        aadhar_number=aadharNumber,
        permanent_address=permanentAddress,
        present_address=presentAddress,

        emergency_name=emergencyName,
        emergency_relation=emergencyRelation,
        emergency_phone=emergencyPhone,

        qualification10=qualification10,
        year10=year10,
        percent10=percent10,
        qualification12=qualification12,
        year12=year12,
        percent12=percent12,
        ug_degree=ugDegree,
        ug_college=ugCollege,
        ug_year=ugYear,
        ug_percent=ugPercent,
        pg_degree=pgDegree,
        pg_college=pgCollege,
        pg_year=pgYear,
        pg_percent=pgPercent,

        total_exp_years=totalExpYears,
        total_exp_months=totalExpMonths,
        career_break=careerBreak,
        career_break_duration=careerBreakDuration,
        career_break_reason=careerBreakReason,

        account_holder_name=accountHolderName,
        bank_name=bankName,
        account_number=accountNumber,
        ifsc_code=ifscCode,
        branch_name=branchName,

        esi_applicable=esiApplicable,
        uan_number=uanNumber,
        pf_number=pfNumber,
        esi_number=esiNumber,

        # DOCUMENTS
        resume=resume_file,
        sslc=sslc_file,
        hsc=hsc_file,
        aadharSelf=aadhar_self_file,
        photo=photo_file,
        aadharFather=aadhar_father_file,
        aadharMother=aadhar_mother_file,
        panSelf=pan_self_file,
        bankPassbookPhoto=bank_passbook_file,

        status="Active"
    )

    db.add(employee)
    db.commit()
    db.refresh(employee)
    
    # ================= SAVE DEPENDENTS =================
    if dependents:
      dependents_data = json.loads(dependents)

    for index, dep in enumerate(dependents_data):

        new_dependent = models.Dependent(
            employee_id=employee.id,
            name=dep.get("name"),
            dob=datetime.strptime(dep.get("dob"), "%Y-%m-%d").date() if dep.get("dob") else None,
            relation=dep.get("relation"),
            aadhar_number=dep.get("aadharNumber"),
            pan_number=dep.get("panNumber"),
        )

        db.add(new_dependent)
        db.commit()
        db.refresh(new_dependent)

        # Save Files Dynamically
        for file_type in ["aadhar", "pan", "photo"]:
            file_field = locals().get(f"dependent_{file_type}_{index}")

            if file_field:
                file_name = save_file(file_field)

                if file_type == "aadhar":
                    new_dependent.aadhar_photo = file_name
                elif file_type == "pan":
                    new_dependent.pan_photo = file_name
                elif file_type == "photo":
                    new_dependent.photo = file_name

        db.commit()

    # SAVE TRAININGS
    if trainings:
        for t in json.loads(trainings):
            db.add(models.Training(
                employee_id=employee.id,
                training_name=t.get("name"),
                institute=t.get("institute"),
                duration=t.get("duration"),
                year=t.get("year"),
                remarks=t.get("remarks")
            ))

    # SAVE EMPLOYMENTS
    if employments:
        for e in json.loads(employments):
            db.add(models.Employment(
                employee_id=employee.id,
                organization=e.get("organization"),
                designation=e.get("designation"),
                period=e.get("period"),
                salary=e.get("salary"),
                nature=e.get("nature"),
                reason=e.get("reason")
            ))

    db.commit()

    return {"message": "Employee saved successfully", "employee_id": employee.id}
# GET ALL EMPLOYEES (HR DASHBOARD)
# =====================================================

@app.get("/employees")
def get_employees(db: Session = Depends(get_db)):

    employees = db.query(models.EmployeeJoining).all()

    return [
        {
            column.name: getattr(emp, column.name)
            for column in emp.__table__.columns
        }
        for emp in employees
    ]


# =====================================================
# GET FULL EMPLOYEE (WITH RELATIONS)
# =====================================================

@app.get("/employee-full/{employee_id}")
def get_full_employee(employee_id: int, db: Session = Depends(get_db)):

    employee = db.query(models.EmployeeJoining).filter(
        models.EmployeeJoining.id == employee_id
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    employee_data = {
        column.name: getattr(employee, column.name)
        for column in employee.__table__.columns
    }

    trainings_data = [
        {
            column.name: getattr(training, column.name)
            for column in training.__table__.columns
        }
        for training in employee.trainings
    ]

    employments_data = [
        {
            column.name: getattr(emp, column.name)
            for column in emp.__table__.columns
        }
        for emp in employee.employments
    ]
    dependents_data = [
    {
        column.name: getattr(dep, column.name)
        for column in dep.__table__.columns
    }
    for dep in employee.dependents
    ]

    return {
        "employee": employee_data,
        "trainings": trainings_data,
        "employments": employments_data,
        "dependents":dependents_data
    }



# =====================================================
# DOWNLOAD EMPLOYEE DOCUMENTS AS ZIP
# =====================================================
@app.get("/download-employee/{employee_id}")
def download_employee(employee_id: int, db: Session = Depends(get_db)):

    employee = db.query(models.EmployeeJoining).filter(
        models.EmployeeJoining.id == employee_id
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    zip_filename = f"employee_{employee_id}.zip"
    zip_path = os.path.join(UPLOAD_FOLDER, zip_filename)

    # Remove old zip if exists
    if os.path.exists(zip_path):
        os.remove(zip_path)

    file_fields = [
        "resume",
        "sslc",
        "hsc",
        "aadharSelf",
        "photo",
        "aadharFather",
        "aadharMother",
        "panSelf",
        "bankPassbookPhoto"
    ]

    files_added = 0

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:

        for field in file_fields:
            file_name = getattr(employee, field, None)

            if file_name:
                file_path = os.path.join(UPLOAD_FOLDER, file_name)

                # Debug print (important)
                print("Checking:", file_path)

                if os.path.exists(file_path):
                    zipf.write(file_path, arcname=file_name)
                    files_added += 1
                else:
                    print("File NOT found:", file_path)

    if files_added == 0:
        raise HTTPException(status_code=400, detail="No documents found for this employee")

    return FileResponse(
        zip_path,
        filename=zip_filename,
        media_type="application/zip"
    )

    employee = db.query(models.EmployeeJoining).filter(
        models.EmployeeJoining.id == employee_id
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    zip_filename = f"employee_{employee_id}.zip"
    zip_path = os.path.join(UPLOAD_FOLDER, zip_filename)

    # 🔹 List ONLY file fields
    file_fields = [
        "resume",
        "sslc",
        "hsc",
        "graduation",
        "post_graduation",
        "aadhar",
        "pan",
        "photo",
        "bank_passbook"
    ]

    with zipfile.ZipFile(zip_path, "w") as zipf:

        for field in file_fields:
            file_name = getattr(employee, field)

            if file_name:
                file_path = os.path.join(UPLOAD_FOLDER, file_name)

                if os.path.exists(file_path):
                    zipf.write(file_path, arcname=file_name)

    return FileResponse(
        zip_path,
        filename=zip_filename,
        media_type="application/zip"
    )

@app.get("/download-employee-excel/{employee_id}")
def download_employee_excel(employee_id: int, db: Session = Depends(get_db)):

    employee = db.query(models.EmployeeJoining)\
        .options(
            joinedload(models.EmployeeJoining.trainings),
            joinedload(models.EmployeeJoining.employments),
            joinedload(models.EmployeeJoining.dependents)
        )\
        .filter(models.EmployeeJoining.id == employee_id)\
        .first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Report"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    row = 1

    # ================= TITLE =================
    ws.merge_cells("A1:AZ1")
    ws["A1"] = "EMPLOYEE MASTER DATA REPORT"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = center
    row += 2

    # ================= COMPLETE BASIC DETAILS =================
    headers = [
        "Employee Code", "Name", "DOB", "DOJ",
        "Gender", "Phone", "Email",
        "Father Name", "Mother Name",
        "Department", "Designation",
        "Blood Group", "Marital Status", "Spouse Name",
        "Aadhar", "PAN",
        "Permanent Address", "Present Address",
        "Emergency Name", "Emergency Relation", "Emergency Phone",
        "10th School", "10th Year", "10th %",
        "12th School", "12th Year", "12th %",
        "UG Degree", "UG College", "UG Year", "UG %",
        "PG Degree", "PG College", "PG Year", "PG %",
        "Total Exp Years", "Total Exp Months",
        "Career Break", "Break Duration", "Break Reason",
        "ESI Applicable", "UAN", "PF", "ESI",
        "Account Holder", "Bank Name", "Account Number",
        "IFSC", "Branch", "Status"
    ]

    data = [
        employee.employee_code,
        employee.name,
        employee.dob,
        employee.doj,
        employee.gender,
        employee.phone,
        employee.email,
        employee.father_name,
        employee.mother_name,
        employee.department,
        employee.designation,
        employee.blood_group,
        employee.marital_status,
        employee.spouse_name,
        employee.aadhar_number,
        employee.pan_number,
        employee.permanent_address,
        employee.present_address,
        employee.emergency_name,
        employee.emergency_relation,
        employee.emergency_phone,
        employee.qualification10,
        employee.year10,
        employee.percent10,
        employee.qualification12,
        employee.year12,
        employee.percent12,
        employee.ug_degree,
        employee.ug_college,
        employee.ug_year,
        employee.ug_percent,
        employee.pg_degree,
        employee.pg_college,
        employee.pg_year,
        employee.pg_percent,
        employee.total_exp_years,
        employee.total_exp_months,
        employee.career_break,
        employee.career_break_duration,
        employee.career_break_reason,
        employee.esi_applicable,
        employee.uan_number,
        employee.pf_number,
        employee.esi_number,
        employee.account_holder_name,
        employee.bank_name,
        employee.account_number,
        employee.ifsc_code,
        employee.branch_name,
        employee.status
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header).font = bold

    row += 1

    for col, value in enumerate(data, 1):
        ws.cell(row=row, column=col, value=value)

    row += 3

    # ================= DEPENDENTS =================
    ws.cell(row=row, column=1, value="DEPENDENT DETAILS").font = bold
    row += 1

    if employee.dependents:
        dep_headers = ["Name", "DOB", "Relation", "Aadhar", "PAN"]

        for col, header in enumerate(dep_headers, 1):
            ws.cell(row=row, column=col, value=header).font = bold

        row += 1

        for dep in employee.dependents:
            ws.cell(row=row, column=1, value=dep.name)
            ws.cell(row=row, column=2, value=dep.dob)
            ws.cell(row=row, column=3, value=dep.relation)
            ws.cell(row=row, column=4, value=dep.aadhar_number)
            ws.cell(row=row, column=5, value=dep.pan_number)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No Dependents")
        row += 2

    row += 2

    # ================= TRAININGS =================
    ws.cell(row=row, column=1, value="TRAININGS").font = bold
    row += 1

    if employee.trainings:
        train_headers = ["Training", "Institute", "Duration", "Year", "Remarks"]

        for col, header in enumerate(train_headers, 1):
            ws.cell(row=row, column=col, value=header).font = bold

        row += 1

        for t in employee.trainings:
            ws.cell(row=row, column=1, value=t.training_name)
            ws.cell(row=row, column=2, value=t.institute)
            ws.cell(row=row, column=3, value=t.duration)
            ws.cell(row=row, column=4, value=t.year)
            ws.cell(row=row, column=5, value=t.remarks)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No Trainings")
        row += 2

    row += 2

    # ================= EMPLOYMENT =================
    ws.cell(row=row, column=1, value="EMPLOYMENT HISTORY").font = bold
    row += 1

    if employee.employments:
        emp_headers = ["Organization", "Designation", "Period", "Salary", "Nature", "Reason"]

        for col, header in enumerate(emp_headers, 1):
            ws.cell(row=row, column=col, value=header).font = bold

        row += 1

        for e in employee.employments:
            ws.cell(row=row, column=1, value=e.organization)
            ws.cell(row=row, column=2, value=e.designation)
            ws.cell(row=row, column=3, value=e.period)
            ws.cell(row=row, column=4, value=e.salary)
            ws.cell(row=row, column=5, value=e.nature)
            ws.cell(row=row, column=6, value=e.reason)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No Employment History")

    # ================= AUTO COLUMN WIDTH =================
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col)

        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    # ================= SAVE =================
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
                f"attachment; filename=employee_{employee.employee_code or employee.id}.xlsx"
        },
    )
    
@app.post("/send-otp")
async def send_otp(data: dict = Body(...), db: Session = Depends(get_db)):
    email = data.get("email")

    if not email:
        return {"success": False, "message": "Email required"}

    # Check if already registered
    existing = db.query(models.EmployeeAuth).filter(
        models.EmployeeAuth.email == email
    ).first()

    if existing and existing.password:
        return {"success": False, "message": "Already registered"}

    if not existing:
        existing = models.EmployeeAuth(
            email=email,
            is_verified=False
        )
        db.add(existing)

    otp = str(random.randint(100000, 999999))
    existing.otp = otp
    existing.otp_expiry = datetime.utcnow() + timedelta(minutes=5)

    db.commit()

    print("OTP:", otp)

    message = MessageSchema(
        subject="OTP Verification",
        recipients=[email],
        body=f"Your OTP is {otp}",
        subtype="plain"
    )

    fm = FastMail(conf)
    await fm.send_message(message)

    return {"success": True}

@app.post("/verify-otp")
def verify_otp(data: dict = Body(...), db: Session = Depends(get_db)):
    email = data.get("email")
    otp = data.get("otp")

    auth = db.query(models.EmployeeAuth).filter(
        models.EmployeeAuth.email == email
    ).first()

    if not auth:
        return {"success": False}

    if auth.otp != otp:
        return {"success": False, "message": "Invalid OTP"}

    if auth.otp_expiry < datetime.utcnow():
        return {"success": False, "message": "OTP Expired"}

    auth.is_verified = True
    db.commit()

    return {"success": True}
@app.post("/set-password")
def set_password(data: dict = Body(...), db: Session = Depends(get_db)):
    email = data.get("email")
    password = data.get("password")

    auth = db.query(models.EmployeeAuth).filter(
        models.EmployeeAuth.email == email,
        models.EmployeeAuth.is_verified == True
    ).first()

    if not auth:
        return {"success": False, "message": "Not verified"}

    auth.password = pwd_context.hash(password)
    db.commit()

    return {"success": True}

@app.post("/employee-login")
def employee_login(data: dict = Body(...), db: Session = Depends(get_db)):
    email = data.get("email")
    password = data.get("password")

    auth = db.query(models.EmployeeAuth).filter(
        models.EmployeeAuth.email == email
    ).first()

    if not auth:
        return {"success": False}

    if not pwd_context.verify(password, auth.password):
        return {"success": False}

    employee = db.query(models.EmployeeJoining).filter(
        models.EmployeeJoining.id == auth.employee_id
    ).first()

    return {
        "success": True,
        "employee": {
            "id": employee.id,
            "name": employee.name,
            "email": employee.email
        }
    }
    
    
@app.delete("/delete-employee/{employee_id}")
def delete_employee(employee_id: int, db: Session = Depends(get_db)):

    employee = db.query(models.EmployeeJoining).filter(
        models.EmployeeJoining.id == employee_id
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Delete related trainings
    db.query(models.Training).filter(
        models.Training.employee_id == employee_id
    ).delete()

    # Delete related employment
    db.query(models.Employment).filter(
        models.Employment.employee_id == employee_id
    ).delete()
    
    db.query(models.Dependent).filter(
    models.Dependent.employee_id == employee_id
    ).delete()

    # Delete main employee
    db.delete(employee)
    db.commit()

    return {"success": True, "message": "Employee deleted successfully"}


    email = data.get("email")
    password = data.get("password")

    employee = db.query(models.EmployeeJoining).filter(
        models.EmployeeJoining.email == email
    ).first()

    if not employee:
        return {"success": False, "message": "Invalid Email or Password"}

    if not employee.is_verified:
        return {"success": False, "message": "Account not verified"}

    if not pwd_context.verify(password, employee.password):
        return {"success": False, "message": "Invalid Email or Password"}

    return {
        "success": True,
        "employee": {
            "id": employee.id,
            "email": employee.email,
            "name": employee.name
        }
    }
    email = data.get("email")
    password = data.get("password")

    if not email or not password:
        return {"success": False, "message": "Email and password required"}

    employee = db.query(models.EmployeeJoining).filter(
        models.EmployeeJoining.email == email
    ).first()

    # Email not found
    if not employee:
        return {"success": False, "message": "Invalid Email or Password"}

    # Account not verified (OTP not completed)
    if not employee.is_verified:
        return {"success": False, "message": "Account not verified"}

    # 🔒 SAFE BYTE TRUNCATION (bcrypt limit fix)
    password_bytes = password.encode("utf-8")[:72]
    password_safe = password_bytes.decode("utf-8", errors="ignore")

    # Verify password
    if not pwd_context.verify(password_safe, employee.password):
        return {"success": False, "message": "Invalid Email or Password"}

    return {
        "success": True,
        "employee": {
            "id": employee.id,
            "email": employee.email,
            "name": employee.name
        }
    }

    employee = db.query(models.EmployeeJoining).filter(
        models.EmployeeJoining.id == employee_id
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Details"

    # ===== Styles =====
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # ===== Column Width =====
    widths = [20, 25, 20, 25, 20, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ===== TITLE ROW =====
    ws.merge_cells("A1:E2")
    ws["A1"] = "Employee Details"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center
    ws["A1"].border = border

    # ===== PHOTO BOX =====
    ws.merge_cells("F1:F6")
    ws["F1"] = "Photo"
    ws["F1"].fill = yellow_fill
    ws["F1"].alignment = center
    ws["F1"].border = border

    # ===== PERSONAL SECTION =====
    row = 3

    fields = [
        ("Name", employee.name),
        ("Father Name", employee.father_name),
        ("Mother Name", employee.mother_name),
        ("Gender", employee.gender),
        ("Department", employee.department),
        ("Designation", employee.designation),
        ("Mobile No", employee.phone),
        ("Email", employee.email),
    ]

    for label, value in fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].fill = yellow_fill
        ws[f"A{row}"].font = bold
        ws[f"A{row}"].border = border

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = grey_fill
        ws[f"B{row}"].border = border

        row += 1

    # ===== ADDRESS SECTION =====
    ws[f"A{row}"] = "Permanent Address"
    ws[f"A{row}"].fill = yellow_fill
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].border = border

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws[f"B{row}"] = employee.permanent_address
    ws[f"B{row}"].fill = grey_fill
    ws[f"B{row}"].border = border
    row += 2

    # ===== BANK SECTION =====
    ws[f"A{row}"] = "Bank Name"
    ws[f"A{row}"].fill = yellow_fill
    ws[f"A{row}"].border = border

    ws[f"B{row}"] = employee.bank_name
    ws[f"B{row}"].fill = grey_fill
    ws[f"B{row}"].border = border

    ws[f"C{row}"] = "Account No"
    ws[f"C{row}"].fill = yellow_fill
    ws[f"C{row}"].border = border

    ws[f"D{row}"] = employee.account_number
    ws[f"D{row}"].fill = grey_fill
    ws[f"D{row}"].border = border

    ws[f"E{row}"] = "IFSC"
    ws[f"E{row}"].fill = yellow_fill
    ws[f"E{row}"].border = border

    ws[f"F{row}"] = employee.ifsc_code
    ws[f"F{row}"].fill = grey_fill
    ws[f"F{row}"].border = border

    row += 2

    # ===== EDUCATION SECTION =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws[f"A{row}"] = "Educational Details"
    ws[f"A{row}"].fill = yellow_fill
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].alignment = center
    ws[f"A{row}"].border = border
    row += 1

    headers = ["Qualification", "Institution", "Year", "Percentage"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col).value = header
        ws.cell(row=row, column=col).fill = yellow_fill
        ws.cell(row=row, column=col).font = bold
        ws.cell(row=row, column=col).border = border

    row += 1

    education = [
        ("10th", employee.qualification10, employee.year10, employee.percent10),
        ("12th", employee.qualification12, employee.year12, employee.percent12),
        ("UG - " + str(employee.ug_degree), employee.ug_college, employee.ug_year, employee.ug_percent),
    ]

    if employee.pg_degree:
        education.append(
            ("PG - " + str(employee.pg_degree), employee.pg_college, employee.pg_year, employee.pg_percent)
        )

    for edu in education:
        for col, value in enumerate(edu, 1):
            ws.cell(row=row, column=col).value = value
            ws.cell(row=row, column=col).fill = grey_fill
            ws.cell(row=row, column=col).border = border
        row += 1

    file_path = f"employee_form_{employee_id}.xlsx"
    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ✅ Correct Model Name
    employee = db.query(models.EmployeeJoining).filter(
        models.EmployeeJoining.id == employee_id
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Database"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    row = 1

    # ================= TITLE =================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="EMPLOYEE DATABASE FORM").font = Font(size=14, bold=True)
    ws.cell(row=row, column=1).alignment = center
    row += 2

    # ================= PERSONAL DETAILS =================
    ws.cell(row=row, column=1, value="Personal Details").font = bold
    row += 1

    personal_fields = [
        ("Name", employee.name),
        ("DOB", str(employee.dob) if employee.dob else ""),
        ("Gender", employee.gender),
        ("Phone", employee.phone),
        ("Email", employee.email),
        ("Department", employee.department),
        ("Designation", employee.designation),
        ("Father Name", employee.father_name),
        ("Mother Name", employee.mother_name),
        ("Permanent Address", employee.permanent_address),
        ("Present Address", employee.present_address),
    ]

    for label, value in personal_fields:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # ================= EDUCATION =================
    ws.cell(row=row, column=1, value="Education Details").font = bold
    row += 1

    headers = ["Level", "Degree", "Institution", "Year", "Percentage"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header).font = bold
    row += 1

    education_data = [
        ("10th", "-", employee.qualification10, employee.year10, employee.percent10),
        ("12th", "-", employee.qualification12, employee.year12, employee.percent12),
        ("UG", employee.ug_degree, employee.ug_college, employee.ug_year, employee.ug_percent),
    ]

    if employee.pg_degree:
        education_data.append(
            ("PG", employee.pg_degree, employee.pg_college, employee.pg_year, employee.pg_percent)
        )

    for edu in education_data:
        for col, value in enumerate(edu, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    row += 2

    # ================= BANK DETAILS =================
    ws.cell(row=row, column=1, value="Bank Details").font = bold
    row += 1

    bank_fields = [
        ("Account Holder", employee.account_holder_name),
        ("Bank Name", employee.bank_name),
        ("Account Number", employee.account_number),
        ("IFSC Code", employee.ifsc_code),
        ("Branch Name", employee.branch_name),
    ]

    for label, value in bank_fields:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # ================= ESI / PF =================
    ws.cell(row=row, column=1, value="ESI / PF Details").font = bold
    row += 1

    esi_fields = [
        ("ESI Applicable", employee.esi_applicable),
        ("UAN Number", employee.uan_number),
        ("PF Number", employee.pf_number),
        ("ESI Number", employee.esi_number),
    ]

    for label, value in esi_fields:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # ================= TRAININGS =================
    ws.cell(row=row, column=1, value="Professional Trainings").font = bold
    row += 1

    if employee.trainings:
        ws.cell(row=row, column=1, value="Training Name").font = bold
        ws.cell(row=row, column=2, value="Institute").font = bold
        ws.cell(row=row, column=3, value="Duration").font = bold
        ws.cell(row=row, column=4, value="Year").font = bold
        row += 1

        for t in employee.trainings:
            ws.cell(row=row, column=1, value=t.training_name)
            ws.cell(row=row, column=2, value=t.institute)
            ws.cell(row=row, column=3, value=t.duration)
            ws.cell(row=row, column=4, value=t.year)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No Trainings")
        row += 1

    row += 2

    # ================= EMPLOYMENT =================
    ws.cell(row=row, column=1, value="Employment History").font = bold
    row += 1

    if employee.employments:
        ws.cell(row=row, column=1, value="Organization").font = bold
        ws.cell(row=row, column=2, value="Designation").font = bold
        ws.cell(row=row, column=3, value="Period").font = bold
        ws.cell(row=row, column=4, value="Salary").font = bold
        ws.cell(row=row, column=5, value="Reason").font = bold
        row += 1

        for e in employee.employments:
            ws.cell(row=row, column=1, value=e.organization)
            ws.cell(row=row, column=2, value=e.designation)
            ws.cell(row=row, column=3, value=e.period)
            ws.cell(row=row, column=4, value=e.salary)
            ws.cell(row=row, column=5, value=e.reason)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No Employment History")

    file_path = f"employee_{employee_id}.xlsx"
    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    employee = db.query(models.Employee).filter(
        models.Employee.id == employee_id
    ).first()

    trainings = db.query(models.Training).filter(
        models.Training.employee_id == employee_id
    ).all()

    employments = db.query(models.Employment).filter(
        models.Employment.employee_id == employee_id
    ).all()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Database"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # ================= TITLE =================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="EMPLOYEE DATABASE FORM").font = Font(size=14, bold=True)
    ws.cell(row=row, column=1).alignment = center
    row += 2

    # ================= PERSONAL DETAILS =================
    ws.cell(row=row, column=1, value="Personal Details").font = bold
    row += 1

    personal_data = [
        ("Name", employee.name),
        ("DOB", employee.dob),
        ("Gender", employee.gender),
        ("Phone", employee.phone),
        ("Email", employee.email),
        ("Department", employee.department),
        ("Designation", employee.designation),
        ("Father Name", employee.father_name),
        ("Mother Name", employee.mother_name),
        ("Permanent Address", employee.permanent_address),
        ("Present Address", employee.present_address),
    ]

    for label, value in personal_data:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # ================= EDUCATION =================
    ws.cell(row=row, column=1, value="Education Details").font = bold
    row += 1

    headers = ["Level", "Degree", "Institution", "Year", "Percentage"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header).font = bold
    row += 1

    education_rows = [
        ("10th", "-", employee.qualification10, employee.year10, employee.percent10),
        ("12th", "-", employee.qualification12, employee.year12, employee.percent12),
        ("UG", employee.ug_degree, employee.ug_college, employee.ug_year, employee.ug_percent),
    ]

    if employee.pg_degree:
        education_rows.append(
            ("PG", employee.pg_degree, employee.pg_college, employee.pg_year, employee.pg_percent)
        )

    for edu in education_rows:
        for col, value in enumerate(edu, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    row += 2

    # ================= BANK DETAILS =================
    ws.cell(row=row, column=1, value="Bank Details").font = bold
    row += 1

    bank_data = [
        ("Account Holder", employee.account_holder_name),
        ("Bank Name", employee.bank_name),
        ("Account Number", employee.account_number),
        ("IFSC", employee.ifsc_code),
        ("Branch", employee.branch_name),
    ]

    for label, value in bank_data:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # ================= TRAININGS =================
    ws.cell(row=row, column=1, value="Professional Trainings").font = bold
    row += 1

    if trainings:
        ws.cell(row=row, column=1, value="Training Name").font = bold
        ws.cell(row=row, column=2, value="Institute").font = bold
        ws.cell(row=row, column=3, value="Duration").font = bold
        ws.cell(row=row, column=4, value="Year").font = bold
        row += 1

        for t in trainings:
            ws.cell(row=row, column=1, value=t.training_name)
            ws.cell(row=row, column=2, value=t.institute)
            ws.cell(row=row, column=3, value=t.duration)
            ws.cell(row=row, column=4, value=t.year)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No Trainings")
        row += 1

    row += 2

    # ================= EMPLOYMENT =================
    ws.cell(row=row, column=1, value="Employment History").font = bold
    row += 1

    if employments:
        ws.cell(row=row, column=1, value="Organization").font = bold
        ws.cell(row=row, column=2, value="Designation").font = bold
        ws.cell(row=row, column=3, value="Period").font = bold
        ws.cell(row=row, column=4, value="Salary").font = bold
        ws.cell(row=row, column=5, value="Reason").font = bold
        row += 1

        for e in employments:
            ws.cell(row=row, column=1, value=e.organization)
            ws.cell(row=row, column=2, value=e.designation)
            ws.cell(row=row, column=3, value=e.period)
            ws.cell(row=row, column=4, value=e.salary)
            ws.cell(row=row, column=5, value=e.reason)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No Employment History")

    file_path = f"employee_{employee_id}.xlsx"
    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename=file_path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # ✅ Correct Model Name
    employee = db.query(models.EmployeeJoining)\
        .filter(models.EmployeeJoining.id == emp_id)\
        .first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Because you already have relationship
    trainings = employee.trainings
    employments = employee.employments

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Details"

    # ================= PERSONAL DETAILS =================
    ws.append(["PERSONAL DETAILS"])
    ws.append(["Name", employee.name])
    ws.append(["DOB", str(employee.dob)])
    ws.append(["Gender", employee.gender])
    ws.append(["Phone", employee.phone])
    ws.append(["Email", employee.email])
    ws.append(["Department", employee.department])
    ws.append(["Designation", employee.designation])
    ws.append(["Father Name", employee.father_name])
    ws.append(["Mother Name", employee.mother_name])
    ws.append(["Permanent Address", employee.permanent_address])
    ws.append(["Present Address", employee.present_address])
    ws.append([])

    # ================= EDUCATION =================
    ws.append(["EDUCATION DETAILS"])
    ws.append(["10th School", employee.qualification10])
    ws.append(["10th Year", employee.year10])
    ws.append(["10th Percentage", employee.percent10])
    ws.append(["12th School", employee.qualification12])
    ws.append(["12th Year", employee.year12])
    ws.append(["12th Percentage", employee.percent12])
    ws.append(["UG College", employee.ug_college])
    ws.append(["UG Year", employee.ug_year])
    ws.append(["UG Percentage", employee.ug_percent])

    if employee.pg_college:
        ws.append(["PG College", employee.pg_college])
        ws.append(["PG Year", employee.pg_year])
        ws.append(["PG Percentage", employee.pg_percent])

    ws.append([])

    # ================= BANK =================
    ws.append(["BANK DETAILS"])
    ws.append(["Account Holder", employee.account_holder_name])
    ws.append(["Bank Name", employee.bank_name])
    ws.append(["Account Number", employee.account_number])
    ws.append(["IFSC", employee.ifsc_code])
    ws.append(["Branch", employee.branch_name])
    ws.append([])

    # ================= ESI/PF =================
    ws.append(["ESI / PF DETAILS"])
    ws.append(["ESI Applicable", employee.esi_applicable])
    ws.append(["UAN", employee.uan_number])
    ws.append(["PF Number", employee.pf_number])
    ws.append(["ESI Number", employee.esi_number])
    ws.append([])

    # ================= TRAININGS =================
    ws.append(["TRAININGS"])
    for t in trainings:
        ws.append([t.training_name, t.institute, t.duration, t.year, t.remarks])
    ws.append([])

    # ================= EMPLOYMENT =================
    ws.append(["EMPLOYMENT HISTORY"])
    for e in employments:
        ws.append([e.organization, e.designation, e.period, e.salary, e.reason])

    # Save to memory
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=employee_{employee.name}.xlsx"
        },
    )

    employee = db.query(models.Employee).filter(models.Employee.id == emp_id).first()
    trainings = db.query(models.Training).filter(models.Training.employee_id == emp_id).all()
    employments = db.query(models.Employment).filter(models.Employment.employee_id == emp_id).all()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Details"

    # ================= PERSONAL DETAILS =================
    ws.append(["PERSONAL DETAILS"])
    ws.append(["Name", employee.name])
    ws.append(["DOB", employee.dob])
    ws.append(["Gender", employee.gender])
    ws.append(["Phone", employee.phone])
    ws.append(["Email", employee.email])
    ws.append(["Department", employee.department])
    ws.append(["Designation", employee.designation])
    ws.append([])

    # ================= EDUCATION =================
    ws.append(["EDUCATION"])
    ws.append(["10th School", employee.qualification10])
    ws.append(["10th Year", employee.year10])
    ws.append(["10th Percentage", employee.percent10])
    ws.append(["12th School", employee.qualification12])
    ws.append(["12th Year", employee.year12])
    ws.append(["12th Percentage", employee.percent12])
    ws.append(["UG College", employee.ug_college])
    ws.append(["UG Year", employee.ug_year])
    ws.append(["UG Percentage", employee.ug_percent])
    ws.append([])

    # ================= BANK =================
    ws.append(["BANK DETAILS"])
    ws.append(["Account Holder", employee.account_holder_name])
    ws.append(["Bank Name", employee.bank_name])
    ws.append(["Account Number", employee.account_number])
    ws.append(["IFSC", employee.ifsc_code])
    ws.append([])

    # ================= TRAININGS =================
    ws.append(["TRAININGS"])
    for t in trainings:
        ws.append([t.training_name, t.institute, t.duration, t.year])

    ws.append([])

    # ================= EMPLOYMENT =================
    ws.append(["EMPLOYMENT HISTORY"])
    for e in employments:
        ws.append([e.organization, e.designation, e.period, e.salary, e.reason])

    # Save file to memory
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=employee_{employee.name}.xlsx"
        },
    )